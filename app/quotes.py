"""Deterministic commercial quotes with immutable evidence and shared approvals."""

import csv
import hashlib
import io
import re
import time
import zipfile
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

from pydantic import ValidationError
from sqlalchemy import select

from app.auth import get_actor, require
from app.db import digest, record, row, rows, transaction, uid
from app.errors import DomainError
from app.quote_contracts import QuoteDraft, QuoteSuggestion, QuoteSuggestionRequest
from app.quote_schema import catalog_versions, catalogs, quote_versions, quotes
from app.schema_v1 import jobs, proposals, runs

HEADERS = ["sku", "name", "unit", "price_without_vat", "vat_percent", "currency"]
CENT = Decimal("0.01")
MAX_CATALOG_BYTES = 2 * 1024 * 1024


def money(value):
    return format(value.quantize(CENT, rounding=ROUND_HALF_UP), ".2f")


def decimal_cell(value, cell, maximum, places):
    try:
        result = Decimal(str(value).strip())
        if not result.is_finite() or result < 0 or result > maximum:
            raise ValueError()
        if result.as_tuple().exponent < -places:
            raise ValueError()
        return result
    except (InvalidOperation, ValueError, TypeError):
        raise DomainError("INVALID_CATALOG_CELL", 422, f"Invalid number in {cell}") from None


def parse_catalog(filename, content):
    if not content or len(content) > MAX_CATALOG_BYTES:
        raise DomainError("CATALOG_SIZE_LIMIT", 413)
    if len(filename) > 200 or "/" in filename or "\\" in filename:
        raise DomainError("UNSUPPORTED_CATALOG", 415)
    extension = Path(filename).suffix.lower()
    try:
        if extension == ".csv":
            stream = io.StringIO(content.decode("utf-8-sig"))
            raw = list(csv.reader(stream, strict=True))
        elif extension == ".xlsx":
            from openpyxl import load_workbook

            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                if sum(info.file_size for info in archive.infolist()) > 20 * 1024 * 1024:
                    raise DomainError("CATALOG_EXPANSION_LIMIT", 413)
                if len(archive.infolist()) > 500:
                    raise DomainError("CATALOG_EXPANSION_LIMIT", 413)
            workbook = load_workbook(
                io.BytesIO(content), read_only=True, data_only=False, keep_links=False
            )
            try:
                if len(workbook.worksheets) != 1:
                    raise DomainError("CATALOG_SINGLE_SHEET_REQUIRED", 422)
                sheet = workbook.worksheets[0]
                # The declared XLSX dimensions are untrusted and can hide real rows.
                sheet.reset_dimensions()
                raw = []
                for cells in sheet.iter_rows():
                    if len(raw) >= 1001:
                        raise DomainError("CATALOG_ROW_LIMIT", 422)
                    if len(cells) > len(HEADERS):
                        raise DomainError("INVALID_CATALOG_COLUMNS", 422)
                    for cell in cells:
                        if cell.data_type == "f" or (
                            isinstance(cell.value, str) and cell.value.startswith("=")
                        ):
                            raise DomainError(
                                "CATALOG_FORMULA_NOT_ALLOWED", 422, f"Formula in {cell.coordinate}"
                            )
                    values = [cell.value for cell in cells]
                    raw.append(values)
            finally:
                workbook.close()
        else:
            raise DomainError("UNSUPPORTED_CATALOG", 415)
    except DomainError:
        raise
    except Exception as exc:
        raise DomainError("INVALID_CATALOG_FILE", 422) from exc
    if not raw or [str(v or "").strip() for v in raw[0]] != HEADERS:
        raise DomainError("INVALID_CATALOG_COLUMNS", 422, "Expected columns: " + ", ".join(HEADERS))
    if not 1 <= len(raw) - 1 <= 1000:
        raise DomainError("CATALOG_ROW_LIMIT", 422)
    result, seen = [], set()
    for number, values in enumerate(raw[1:], 2):
        if len(values) != len(HEADERS):
            raise DomainError("INVALID_CATALOG_COLUMNS", 422, f"Invalid row {number}")
        for index, value in enumerate(values):
            if isinstance(value, str) and value.strip().startswith(("=", "+", "@")):
                raise DomainError(
                    "CATALOG_FORMULA_NOT_ALLOWED", 422, f"Formula in {chr(65 + index)}{number}"
                )
        item = dict(zip(HEADERS, values, strict=True))
        for key, limit in (("sku", 80), ("name", 200), ("unit", 40)):
            item[key] = str(item[key] or "").strip()
            if not item[key] or len(item[key]) > limit or any(ord(c) < 32 for c in item[key]):
                raise DomainError("INVALID_CATALOG_CELL", 422, f"Invalid {key} in row {number}")
        if item["sku"] in seen:
            raise DomainError("DUPLICATE_SKU", 422, f"Duplicate SKU in row {number}")
        seen.add(item["sku"])
        if item["currency"] != "RUB":
            raise DomainError(
                "CATALOG_CURRENCY_RUB_REQUIRED", 422, f"Currency in F{number} must be RUB"
            )
        item["price_without_vat"] = money(
            decimal_cell(item["price_without_vat"], f"D{number}", Decimal("1000000000"), 2)
        )
        item["vat_percent"] = str(
            decimal_cell(item["vat_percent"], f"E{number}", Decimal("100"), 2)
        )
        item["source_row"] = number
        result.append(item)
    return result


def calculate(lines, catalog):
    products = {item["sku"]: item for item in catalog["rows"]}
    calculated = []
    net_total, vat_total = Decimal(0), Decimal(0)
    for selection in lines:
        item = products.get(selection.sku)
        if item is None:
            raise DomainError(
                "UNKNOWN_SKU", 422, "A selected SKU is absent from the chosen catalog version"
            )
        gross_net = Decimal(item["price_without_vat"]) * selection.quantity
        net = (gross_net * (1 - selection.discount_percent / 100)).quantize(
            CENT, rounding=ROUND_HALF_UP
        )
        vat = (net * Decimal(item["vat_percent"]) / 100).quantize(CENT, rounding=ROUND_HALF_UP)
        net_total += net
        vat_total += vat
        calculated.append(
            {
                **item,
                "quantity": str(selection.quantity),
                "discount_percent": str(selection.discount_percent),
                "net": money(net),
                "vat": money(vat),
                "total": money(net + vat),
                "evidence": {
                    "catalog_version_id": catalog["id"],
                    "catalog_version": catalog["version"],
                    "content_hash": catalog["content_hash"],
                    "source_row": item["source_row"],
                },
                "formula": "net = ROUND_HALF_UP(price_without_vat * quantity * (1 - discount_percent / 100), 2); vat = ROUND_HALF_UP(net * vat_percent / 100, 2); total = net + vat",
            }
        )
    return {
        "lines": calculated,
        "net": money(net_total),
        "vat": money(vat_total),
        "total": money(net_total + vat_total),
        "currency": "RUB",
        "rounding": "ROUND_HALF_UP per line to 0.01",
    }


class Quotes:
    def __init__(self, engine, settings, knowledge, clock=time.time):
        self.engine, self.settings, self.knowledge, self.clock = engine, settings, knowledge, clock

    def _actor(self, conn, actor):
        current = get_actor(conn, actor.id)
        if current.organization_id != actor.organization_id:
            raise DomainError("NOT_FOUND", 404)
        require(current, "owner", "manager")
        return current

    def catalog(self, actor, version_id, conn=None):
        if conn is None:
            with self.engine.connect() as connection:
                return self.catalog(actor, version_id, connection)
        actor = self._actor(conn, actor)
        version = row(conn, select(catalog_versions).where(catalog_versions.c.id == version_id))
        catalog = (
            row(
                conn,
                select(catalogs).where(
                    catalogs.c.id == version["catalog_id"],
                    catalogs.c.organization_id == actor.organization_id,
                ),
            )
            if version
            else None
        )
        if not catalog or catalog["revoked"] or actor.role not in catalog["roles"]:
            raise DomainError("NOT_FOUND", 404)
        if digest(version["rows"]) != version["content_hash"]:
            raise DomainError("CATALOG_HASH_MISMATCH", 409)
        return {
            **version,
            "name": catalog["name"],
            "current": version["version"] == catalog["current_version"],
        }

    def list_catalogs(self, actor):
        with self.engine.connect() as conn:
            actor = self._actor(conn, actor)
            result = []
            for catalog in rows(
                conn,
                select(catalogs)
                .where(
                    catalogs.c.organization_id == actor.organization_id,
                    catalogs.c.revoked.is_(False),
                )
                .order_by(catalogs.c.name),
            ):
                if actor.role in catalog["roles"]:
                    version = row(
                        conn,
                        select(catalog_versions).where(
                            catalog_versions.c.catalog_id == catalog["id"],
                            catalog_versions.c.version == catalog["current_version"],
                        ),
                    )
                    result.append(
                        {
                            **catalog,
                            "version_id": version["id"],
                            "content_hash": version["content_hash"],
                            "row_count": len(version["rows"]),
                        }
                    )
            return result

    def import_catalog(self, actor, filename, content, correlation_id, catalog_id=None):
        require(actor, "owner", "manager")
        parsed = parse_catalog(filename, content)
        with transaction(self.engine) as conn:
            actor = self._actor(conn, actor)
            catalog = (
                row(
                    conn,
                    select(catalogs).where(
                        catalogs.c.id == catalog_id,
                        catalogs.c.organization_id == actor.organization_id,
                    ),
                )
                if catalog_id
                else None
            )
            if catalog_id and (
                not catalog or catalog["revoked"] or actor.role not in catalog["roles"]
            ):
                raise DomainError("NOT_FOUND", 404)
            if not catalog:
                catalog = {
                    "id": uid(),
                    "organization_id": actor.organization_id,
                    "name": filename,
                    "roles": ["owner", "manager"],
                    "revoked": False,
                    "current_version": 0,
                }
                conn.execute(catalogs.insert().values(**catalog))
            previous = row(
                conn,
                select(catalog_versions).where(
                    catalog_versions.c.catalog_id == catalog["id"],
                    catalog_versions.c.version == catalog["current_version"],
                ),
            )
            hashed = digest(parsed)
            if previous and previous["content_hash"] == hashed:
                return {**previous, "replayed": True}
            version = {
                "id": uid(),
                "catalog_id": catalog["id"],
                "version": catalog["current_version"] + 1,
                "rows": parsed,
                "content_hash": hashed,
                "source_hash": hashlib.sha256(content).hexdigest(),
                "created_at": self.clock(),
            }
            conn.execute(catalog_versions.insert().values(**version))
            conn.execute(
                catalogs.update()
                .where(catalogs.c.id == catalog["id"])
                .values(current_version=version["version"])
            )
            record(
                conn,
                actor,
                "catalog_import",
                catalog["id"],
                "succeeded",
                correlation_id,
                {"version": version["version"], "content_hash": hashed},
                now=self.clock(),
            )
            return {**version, "replayed": False}

    def update_catalog_acl(self, actor, catalog_id, acl, correlation_id):
        with transaction(self.engine) as conn:
            actor = self._actor(conn, actor)
            require(actor, "owner")
            catalog = row(
                conn,
                select(catalogs).where(
                    catalogs.c.id == catalog_id, catalogs.c.organization_id == actor.organization_id
                ),
            )
            if not catalog:
                raise DomainError("NOT_FOUND", 404)
            conn.execute(
                catalogs.update()
                .where(catalogs.c.id == catalog_id)
                .values(roles=acl.roles, revoked=acl.revoked)
            )
            record(
                conn,
                actor,
                "catalog_acl",
                catalog_id,
                "succeeded",
                correlation_id,
                {"roles": acl.roles, "revoked": acl.revoked},
                now=self.clock(),
            )
        return {"id": catalog_id, **acl.model_dump()}

    def _source(self, actor, payload, conn):
        actor = self._actor(conn, actor)
        if not payload.source_document_id:
            return None
        doc = self.knowledge.get_document(
            actor, payload.source_document_id, payload.source_document_version, conn=conn
        )
        return {
            "document_id": doc["id"],
            "version": payload.source_document_version,
            "content_hash": doc["source"]["content_hash"],
            "name": doc["name"],
        }

    def validate_suggestion(self, actor, payload, conn=None):
        request = QuoteSuggestionRequest.model_validate(payload)
        if conn is None:
            with self.engine.connect() as connection:
                return self.validate_suggestion(actor, payload, connection)
        self.catalog(actor, request.catalog_version_id, conn)
        self._source(actor, request, conn)
        return request

    def suggest(self, actor, payload, provider):
        with self.engine.connect() as conn:
            request = self.validate_suggestion(actor, payload, conn)
            catalog = self.catalog(actor, request.catalog_version_id, conn)
            text = request.request_text
            if request.source_document_id:
                source = self.knowledge.get_document(
                    actor, request.source_document_id, request.source_document_version, conn=conn
                )
                text += "\n" + source["content"]
        if len(text) > 12000:
            raise DomainError("REQUEST_TOO_LARGE_FOR_SUGGESTION", 422)
        if provider.engine == "deterministic_demo":
            suggestions = []
            for item in catalog["rows"]:
                match = re.search(
                    r"(?<![\w-])"
                    + re.escape(item["sku"])
                    + r"\s*(?:[xх×:=]\s*|количество\s*)(\d+(?:[.,]\d{1,3})?)(?![\d.,])",
                    text,
                    re.IGNORECASE,
                )
                if match:
                    suggestions.append({"sku": item["sku"], "quantity": match[1].replace(",", ".")})
            answer = {
                "lines": suggestions,
                "unresolved": [
                    "Review every requested position: demo matches only explicit SKU × quantity; other text is not interpreted."
                ],
                "accompanying_text": "Проект коммерческого предложения. Состав и количество требуют проверки.",
            }
        elif hasattr(provider, "suggest_quote"):
            answer = provider.suggest_quote(text, catalog["rows"])
        else:
            raise DomainError("QUOTE_SUGGESTIONS_UNAVAILABLE", 503)
        try:
            suggestion = QuoteSuggestion.model_validate(answer)
        except ValidationError as exc:
            raise DomainError("INVALID_QUOTE_SUGGESTION", 422) from exc
        if len({line.sku for line in suggestion.lines}) != len(suggestion.lines):
            raise DomainError("INVALID_QUOTE_SUGGESTION", 422)
        calculate(
            suggestion.lines, catalog
        )  # Unknown model-generated SKUs must never survive validation.
        self.validate_suggestion(actor, payload)
        return {
            **suggestion.model_dump(mode="json"),
            "catalog_version_id": catalog["id"],
            "catalog_hash": catalog["content_hash"],
            "engine": provider.engine,
            "model_id": provider.model_id,
            "synthetic": provider.engine == "deterministic_demo",
            "requires_review": True,
        }

    def propose_lines(
        self,
        actor,
        catalog_version_id,
        request_text,
        provider,
        source_document_id=None,
        source_document_version=None,
    ):
        return self.suggest(
            actor,
            {
                "catalog_version_id": catalog_version_id,
                "request_text": request_text,
                "source_document_id": source_document_id,
                "source_document_version": source_document_version,
            },
            provider,
        )

    def _quote(self, actor, quote_id, conn):
        actor = self._actor(conn, actor)
        result = row(
            conn,
            select(quotes).where(
                quotes.c.id == quote_id,
                quotes.c.organization_id == actor.organization_id,
                quotes.c.actor_id == actor.id,
            ),
        )
        if not result:
            raise DomainError("NOT_FOUND", 404)
        return result

    def _validate_snapshot(self, conn, actor, version):
        snapshot = version["snapshot"]
        if digest(snapshot) != version["content_hash"]:
            raise DomainError("QUOTE_HASH_MISMATCH", 409)
        self.catalog(actor, snapshot["input"]["catalog_version_id"], conn)
        payload = QuoteDraft.model_validate(snapshot["input"])
        source = self._source(actor, payload, conn)
        if source != snapshot["source"]:
            raise DomainError("SOURCE_HASH_MISMATCH", 409)
        return snapshot

    def get(self, actor, quote_id, version=None, conn=None):
        if conn is None:
            with self.engine.connect() as connection:
                return self.get(actor, quote_id, version, connection)
        quote = self._quote(actor, quote_id, conn)
        revision = row(
            conn,
            select(quote_versions).where(
                quote_versions.c.quote_id == quote_id,
                quote_versions.c.version == (version or quote["current_version"]),
            ),
        )
        if not revision:
            raise DomainError("NOT_FOUND", 404)
        self._validate_snapshot(conn, actor, revision)
        result = {**quote, "revision": revision}
        if revision["run_id"]:
            result["run"] = row(conn, select(runs).where(runs.c.id == revision["run_id"]))
            if result["run"] and result["run"]["state"] in {"rejected", "failed", "superseded"}:
                result["revision"]["status"] = result["run"]["state"]
            result["proposal"] = row(
                conn,
                select(proposals).where(
                    proposals.c.run_id == revision["run_id"],
                    proposals.c.version == revision["version"],
                ),
            )
        return result

    def list_quotes(self, actor):
        with self.engine.connect() as conn:
            actor = self._actor(conn, actor)
            result = []
            for quote in rows(
                conn,
                select(quotes)
                .where(
                    quotes.c.organization_id == actor.organization_id, quotes.c.actor_id == actor.id
                )
                .order_by(quotes.c.updated_at.desc())
                .limit(100),
            ):
                try:
                    full = self.get(actor, quote["id"], conn=conn)
                    result.append(
                        {
                            **quote,
                            "title": full["revision"]["snapshot"]["input"]["title"],
                            "total": full["revision"]["snapshot"]["calculation"]["total"],
                            "status": full["revision"]["status"],
                        }
                    )
                except DomainError as exc:
                    if exc.status not in {403, 404}:
                        raise
            return result

    def save(self, actor, payload, correlation_id, quote_id=None, expected_version=None):
        payload = QuoteDraft.model_validate(payload)
        with transaction(self.engine) as conn:
            actor = self._actor(conn, actor)
            catalog = self.catalog(actor, payload.catalog_version_id, conn)
            source = self._source(actor, payload, conn)
            quote = self._quote(actor, quote_id, conn) if quote_id else None
            if quote and quote["current_version"] != expected_version:
                raise DomainError("VERSION_CONFLICT", 409)
            now = self.clock()
            if not quote:
                quote = {
                    "id": uid(),
                    "organization_id": actor.organization_id,
                    "actor_id": actor.id,
                    "current_version": 0,
                    "created_at": now,
                    "updated_at": now,
                }
                conn.execute(quotes.insert().values(**quote))
            else:
                previous = row(
                    conn,
                    select(quote_versions).where(
                        quote_versions.c.quote_id == quote["id"],
                        quote_versions.c.version == quote["current_version"],
                    ),
                )
                if previous["run_id"]:
                    run_id = previous["run_id"]
                    conn.execute(
                        proposals.update()
                        .where(proposals.c.run_id == run_id, proposals.c.status != "executed")
                        .values(status="superseded")
                    )
                    conn.execute(
                        jobs.update()
                        .where(jobs.c.run_id == run_id, jobs.c.status.in_(["queued", "running"]))
                        .values(status="cancelled", lease_token=None)
                    )
                    conn.execute(
                        runs.update()
                        .where(runs.c.id == run_id, runs.c.state != "completed")
                        .values(state="superseded", updated_at=now)
                    )
            number = quote["current_version"] + 1
            snapshot = {
                "input": payload.model_dump(mode="json"),
                "source": source,
                "calculation": calculate(payload.lines, catalog),
                "catalog_name": catalog["name"],
                "catalog_version": catalog["version"],
            }
            revision = {
                "id": uid(),
                "quote_id": quote["id"],
                "version": number,
                "snapshot": snapshot,
                "content_hash": digest(snapshot),
                "status": "draft",
                "run_id": None,
                "created_at": now,
            }
            conn.execute(quote_versions.insert().values(**revision))
            conn.execute(
                quotes.update()
                .where(quotes.c.id == quote["id"])
                .values(current_version=number, updated_at=now)
            )
            record(
                conn,
                actor,
                "quote_saved",
                quote["id"],
                "succeeded",
                correlation_id,
                {"version": number, "content_hash": revision["content_hash"]},
                now=now,
            )
            return self.get(actor, quote["id"], conn=conn)

    def _task(self, conn, actor, task):
        if not task:
            raise DomainError("QUOTE_TASK_REQUIRED", 422)
        assignee = get_actor(conn, task["assignee_id"])
        if assignee.organization_id != actor.organization_id or assignee.team_id != task["team_id"]:
            raise DomainError("INVALID_ASSIGNEE", 422)
        return task

    def propose(self, actor, quote_id, version, key, correlation_id):
        if not key or len(key) > 100:
            raise DomainError("IDEMPOTENCY_KEY_REQUIRED", 422)
        with transaction(self.engine) as conn:
            actor = self._actor(conn, actor)
            full = self.get(actor, quote_id, conn=conn)
            revision = full["revision"]
            if version != full["current_version"]:
                raise DomainError("VERSION_CONFLICT", 409)
            hashed = digest({"kind": "quote", "quote_id": quote_id, "version": version})
            previous = row(
                conn, select(runs).where(runs.c.actor_id == actor.id, runs.c.idempotency_key == key)
            )
            if previous:
                if previous["input_hash"] != hashed:
                    raise DomainError("IDEMPOTENCY_CONFLICT", 409)
                return self.get(actor, quote_id, conn=conn)
            if revision["run_id"]:
                raise DomainError("QUOTE_ALREADY_PROPOSED", 409)
            if not self.catalog(actor, revision["snapshot"]["input"]["catalog_version_id"], conn)[
                "current"
            ]:
                raise DomainError("STALE_CATALOG", 409)
            self._task(conn, actor, revision["snapshot"]["input"]["task"])
            now, run_id = self.clock(), uid()
            payload = {
                "quote_id": quote_id,
                "quote_version": version,
                "quote_hash": revision["content_hash"],
                "snapshot": revision["snapshot"],
            }
            conn.execute(
                runs.insert().values(
                    id=run_id,
                    organization_id=actor.organization_id,
                    actor_id=actor.id,
                    type="quote",
                    state="awaiting_approval",
                    version=version,
                    input={"quote_id": quote_id, "version": version},
                    result={"quote_id": quote_id, "version": version},
                    created_at=now,
                    updated_at=now,
                    correlation_id=correlation_id,
                    idempotency_key=key,
                    input_hash=hashed,
                )
            )
            conn.execute(
                proposals.insert().values(
                    id=uid(),
                    run_id=run_id,
                    version=version,
                    payload=payload,
                    payload_hash=digest(payload),
                    expires_at=now + 3600,
                    status="pending",
                )
            )
            conn.execute(
                quote_versions.update()
                .where(quote_versions.c.id == revision["id"])
                .values(status="awaiting_approval", run_id=run_id)
            )
            record(
                conn,
                actor,
                "quote_proposed",
                quote_id,
                "proposed",
                correlation_id,
                {"version": version, "content_hash": revision["content_hash"]},
                now=now,
            )
            return self.get(actor, quote_id, conn=conn)

    def validate_approval(self, conn, actor, run, proposal):
        full = self.get(actor, run["input"]["quote_id"], conn=conn)
        revision = full["revision"]
        catalog = self.catalog(actor, revision["snapshot"]["input"]["catalog_version_id"], conn)
        if not catalog["current"]:
            raise DomainError("STALE_CATALOG", 409)
        if (
            proposal["payload"]["quote_id"] != full["id"]
            or proposal["payload"]["quote_version"] != run["version"]
            or revision["version"] != run["version"]
            or revision["run_id"] != run["id"]
            or proposal["payload"]["quote_hash"] != revision["content_hash"]
            or proposal["payload"]["snapshot"] != revision["snapshot"]
        ):
            raise DomainError("VERSION_CONFLICT", 409)
        return self._task(conn, actor, revision["snapshot"]["input"]["task"])

    validate_proposal = validate_approval

    def execute_approved(self, conn, actor, run, proposal):
        task = self.validate_approval(conn, actor, run, proposal)
        conn.execute(
            quote_versions.update()
            .where(
                quote_versions.c.quote_id == run["input"]["quote_id"],
                quote_versions.c.version == run["version"],
            )
            .values(status="approved")
        )
        record(
            conn,
            actor,
            "quote_approved",
            run["input"]["quote_id"],
            "succeeded",
            run["correlation_id"],
            {"version": run["version"]},
            now=self.clock(),
        )
        return task

    def export(self, actor, quote_id, format_name, version=None):
        from app.quote_exports import export_quote

        full = self.get(actor, quote_id, version)
        output = export_quote(full, format_name)
        self.get(actor, quote_id, version)  # Recheck source/catalog access after serialization too.
        return output

    def validate_run_access(self, conn, actor, run):
        if run["type"] == "quote_suggestion":
            self.validate_suggestion(actor, run["input"], conn)
        elif run["type"] == "quote":
            self.get(actor, run["input"]["quote_id"], run["version"], conn=conn)

    def list_catalog_versions(self, actor, catalog_id):
        with self.engine.connect() as conn:
            actor = self._actor(conn, actor)
            versions = rows(
                conn,
                select(catalog_versions)
                .where(catalog_versions.c.catalog_id == catalog_id)
                .order_by(catalog_versions.c.version.desc()),
            )
            if not versions:
                raise DomainError("NOT_FOUND", 404)
            return [self.catalog(actor, version["id"], conn) for version in versions]


def catalog_template(format_name="csv"):
    if format_name == "csv":
        return (",".join(HEADERS) + "\n").encode("utf-8")
    if format_name == "xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catalog"
        sheet.append(HEADERS)
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = max(
                len(str(col[0].value)) + 3, 16
            )
        result = io.BytesIO()
        workbook.save(result)
        return result.getvalue()
    raise DomainError("UNSUPPORTED_CATALOG", 415)
